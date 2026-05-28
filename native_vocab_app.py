from __future__ import annotations

import random
import sqlite3
import sys
import os
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk

from openpyxl import Workbook, load_workbook


HEADERS = [
    "출처",
    "단어",
    "뜻",
    "예문",
    "예문뜻",
    "비고",
    "정답횟수",
    "오답횟수",
    "총시도",
    "최근결과",
    "최근테스트일",
    "연속정답",
    "암기상태",
]

NUMERIC_HEADERS = {"정답횟수", "오답횟수", "총시도", "연속정답"}


@dataclass
class Word:
    row: int
    source: str
    word: str
    meaning: str
    example: str
    example_meaning: str
    note: str
    correct_count: int
    wrong_count: int
    total_count: int
    last_result: str
    last_tested_at: str
    streak_correct: int
    memory_state: str

    @property
    def accuracy(self) -> str:
        if self.total_count == 0:
            return "0.0%"
        return f"{self.correct_count / self.total_count * 100:.1f}%"


class WorkbookStore:
    def __init__(self, path: Path, excel_path: Path | None = None) -> None:
        self.path = path
        self.excel_path = excel_path or path.with_suffix(".xlsx")
        self.ensure_file()
        self.import_excel_if_empty()

    def connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        return connection

    @contextmanager
    def transaction(self):
        connection = self.connect()
        try:
            yield connection
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    def ensure_file(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.transaction() as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS words (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT NOT NULL DEFAULT '',
                    word TEXT NOT NULL,
                    meaning TEXT NOT NULL,
                    example TEXT NOT NULL DEFAULT '',
                    example_meaning TEXT NOT NULL DEFAULT '',
                    note TEXT NOT NULL DEFAULT '',
                    correct_count INTEGER NOT NULL DEFAULT 0,
                    wrong_count INTEGER NOT NULL DEFAULT 0,
                    total_count INTEGER NOT NULL DEFAULT 0,
                    last_result TEXT NOT NULL DEFAULT '',
                    last_tested_at TEXT NOT NULL DEFAULT '',
                    streak_correct INTEGER NOT NULL DEFAULT 0,
                    memory_state TEXT NOT NULL DEFAULT '미학습',
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(source, word, meaning)
                )
                """
            )
            connection.execute("CREATE INDEX IF NOT EXISTS idx_words_source ON words(source)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_words_wrong_count ON words(wrong_count)")
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS answer_attempts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    word_id INTEGER NOT NULL,
                    mode TEXT NOT NULL DEFAULT '',
                    selected_answer TEXT NOT NULL DEFAULT '',
                    correct_answer TEXT NOT NULL DEFAULT '',
                    is_correct INTEGER NOT NULL,
                    attempted_at TEXT NOT NULL,
                    attempted_date TEXT NOT NULL,
                    FOREIGN KEY(word_id) REFERENCES words(id)
                )
                """
            )
            connection.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_answer_attempts_date_correct
                ON answer_attempts(attempted_date, is_correct)
                """
            )
            connection.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_answer_attempts_word_date
                ON answer_attempts(word_id, attempted_date)
                """
            )

    def import_excel_if_empty(self) -> None:
        if not self.excel_path.exists():
            return
        with self.transaction() as connection:
            count = connection.execute("SELECT COUNT(*) FROM words").fetchone()[0]
        if count == 0:
            self.import_excel()

    @staticmethod
    def columns(sheet) -> dict[str, int]:
        return {
            str(sheet.cell(1, col).value): col
            for col in range(1, sheet.max_column + 1)
            if sheet.cell(1, col).value
        }

    @staticmethod
    def text(value) -> str:
        return "" if value is None else str(value).strip()

    @staticmethod
    def number(value) -> int:
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    def word_from_row(self, row: sqlite3.Row) -> Word:
        return Word(
            row=int(row["id"]),
            source=self.text(row["source"]),
            word=self.text(row["word"]),
            meaning=self.text(row["meaning"]),
            example=self.text(row["example"]),
            example_meaning=self.text(row["example_meaning"]),
            note=self.text(row["note"]),
            correct_count=self.number(row["correct_count"]),
            wrong_count=self.number(row["wrong_count"]),
            total_count=self.number(row["total_count"]),
            last_result=self.text(row["last_result"]),
            last_tested_at=self.text(row["last_tested_at"]),
            streak_correct=self.number(row["streak_correct"]),
            memory_state=self.text(row["memory_state"]) or "미학습",
        )

    def list_words(self) -> list[Word]:
        with self.transaction() as connection:
            rows = connection.execute("SELECT * FROM words ORDER BY source, id").fetchall()
        return [self.word_from_row(row) for row in rows]

    @staticmethod
    def target_date(days_ago: int) -> str:
        days = max(0, int(days_ago))
        return (date.today() - timedelta(days=days)).isoformat()

    def list_wrong_words_by_date(self, days_ago: int) -> list[Word]:
        attempted_date = self.target_date(days_ago)
        with self.transaction() as connection:
            rows = connection.execute(
                """
                SELECT words.*
                FROM words
                JOIN (
                    SELECT word_id, COUNT(*) AS wrong_attempts, MAX(attempted_at) AS last_wrong_at
                    FROM answer_attempts
                    WHERE attempted_date = ? AND is_correct = 0
                    GROUP BY word_id
                ) daily_wrong ON daily_wrong.word_id = words.id
                ORDER BY daily_wrong.wrong_attempts DESC, daily_wrong.last_wrong_at DESC, words.id
                """,
                (attempted_date,),
            ).fetchall()
        return [self.word_from_row(row) for row in rows]

    def wrong_dates(self, limit: int = 30) -> list[dict[str, int | str]]:
        max_rows = max(1, int(limit))
        with self.transaction() as connection:
            rows = connection.execute(
                """
                SELECT attempted_date, COUNT(*) AS attempts, COUNT(DISTINCT word_id) AS words
                FROM answer_attempts
                WHERE is_correct = 0
                GROUP BY attempted_date
                ORDER BY attempted_date DESC
                LIMIT ?
                """,
                (max_rows,),
            ).fetchall()
        today = date.today()
        return [
            {
                "date": str(row["attempted_date"]),
                "daysAgo": max(0, (today - date.fromisoformat(str(row["attempted_date"]))).days),
                "attempts": int(row["attempts"]),
                "words": int(row["words"]),
            }
            for row in rows
        ]

    def save_result(
        self,
        word: Word,
        is_correct: bool,
        selected_answer: str = "",
        correct_answer: str = "",
        mode: str = "",
    ) -> None:
        correct = word.correct_count + (1 if is_correct else 0)
        wrong = word.wrong_count + (0 if is_correct else 1)
        total = word.total_count + 1
        streak = word.streak_correct + 1 if is_correct else 0
        result = "정답" if is_correct else "오답"
        state = self.memory_state(correct, wrong, total, streak)
        now = datetime.now()
        tested_at = now.strftime("%Y-%m-%d %H:%M")
        attempted_at = now.strftime("%Y-%m-%d %H:%M:%S")

        with self.transaction() as connection:
            connection.execute(
                """
                UPDATE words
                SET correct_count = ?,
                    wrong_count = ?,
                    total_count = ?,
                    last_result = ?,
                    last_tested_at = ?,
                    streak_correct = ?,
                    memory_state = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (correct, wrong, total, result, tested_at, streak, state, word.row),
            )
            connection.execute(
                """
                INSERT INTO answer_attempts (
                    word_id, mode, selected_answer, correct_answer,
                    is_correct, attempted_at, attempted_date
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    word.row,
                    mode,
                    selected_answer,
                    correct_answer,
                    1 if is_correct else 0,
                    attempted_at,
                    now.date().isoformat(),
                ),
            )

    def import_excel(self) -> dict[str, int | str]:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {self.excel_path}")

        workbook = load_workbook(self.excel_path)
        added = 0
        updated = 0
        skipped = 0
        try:
            sheet = workbook.active
            columns = self.columns(sheet)
            required = ["단어", "뜻"]
            missing = [header for header in required if header not in columns]
            if missing:
                raise ValueError(f"Excel에 필요한 열이 없습니다: {', '.join(missing)}")

            with self.transaction() as connection:
                for row in range(2, sheet.max_row + 1):
                    source = self.text(sheet.cell(row, columns.get("출처", 0)).value) if "출처" in columns else ""
                    word = self.text(sheet.cell(row, columns["단어"]).value)
                    meaning = self.text(sheet.cell(row, columns["뜻"]).value)
                    if not word or not meaning:
                        skipped += 1
                        continue

                    values = {
                        "source": source,
                        "word": word,
                        "meaning": meaning,
                        "example": self.text(sheet.cell(row, columns.get("예문", 0)).value) if "예문" in columns else "",
                        "example_meaning": self.text(sheet.cell(row, columns.get("예문뜻", 0)).value)
                        if "예문뜻" in columns
                        else "",
                        "note": self.text(sheet.cell(row, columns.get("비고", 0)).value) if "비고" in columns else "",
                        "correct_count": self.number(sheet.cell(row, columns.get("정답횟수", 0)).value)
                        if "정답횟수" in columns
                        else 0,
                        "wrong_count": self.number(sheet.cell(row, columns.get("오답횟수", 0)).value)
                        if "오답횟수" in columns
                        else 0,
                        "total_count": self.number(sheet.cell(row, columns.get("총시도", 0)).value)
                        if "총시도" in columns
                        else 0,
                        "last_result": self.text(sheet.cell(row, columns.get("최근결과", 0)).value)
                        if "최근결과" in columns
                        else "",
                        "last_tested_at": self.text(sheet.cell(row, columns.get("최근테스트일", 0)).value)
                        if "최근테스트일" in columns
                        else "",
                        "streak_correct": self.number(sheet.cell(row, columns.get("연속정답", 0)).value)
                        if "연속정답" in columns
                        else 0,
                        "memory_state": self.text(sheet.cell(row, columns.get("암기상태", 0)).value)
                        if "암기상태" in columns
                        else "미학습",
                    }
                    if not values["memory_state"]:
                        values["memory_state"] = "미학습"

                    exists = connection.execute(
                        """
                        SELECT 1
                        FROM words
                        WHERE source = ? AND word = ? AND meaning = ?
                        """,
                        (source, word, meaning),
                    ).fetchone()
                    cursor = connection.execute(
                        """
                        INSERT INTO words (
                            source, word, meaning, example, example_meaning, note,
                            correct_count, wrong_count, total_count, last_result,
                            last_tested_at, streak_correct, memory_state
                        )
                        VALUES (
                            :source, :word, :meaning, :example, :example_meaning, :note,
                            :correct_count, :wrong_count, :total_count, :last_result,
                            :last_tested_at, :streak_correct, :memory_state
                        )
                        ON CONFLICT(source, word, meaning) DO UPDATE SET
                            example = excluded.example,
                            example_meaning = excluded.example_meaning,
                            note = excluded.note,
                            updated_at = CURRENT_TIMESTAMP
                        """,
                        values,
                    )
                    if exists:
                        updated += 1
                    else:
                        added += 1
        finally:
            workbook.close()

        return {"added": added, "updated": updated, "skipped": skipped, "path": str(self.excel_path)}

    def export_excel(self) -> dict[str, int | str]:
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = "Words"
            sheet.append(HEADERS)
            for word in self.list_words():
                sheet.append(
                    [
                        word.source,
                        word.word,
                        word.meaning,
                        word.example,
                        word.example_meaning,
                        word.note,
                        word.correct_count,
                        word.wrong_count,
                        word.total_count,
                        word.last_result,
                        word.last_tested_at,
                        word.streak_correct,
                        word.memory_state,
                    ]
                )
            try:
                workbook.save(self.excel_path)
            except PermissionError as exc:
                raise PermissionError(f"Excel 파일을 저장할 수 없습니다. 파일을 닫고 다시 시도하세요: {self.excel_path}") from exc
        finally:
            workbook.close()
        return {"exported": len(self.list_words()), "path": str(self.excel_path)}

    @staticmethod
    def memory_state(correct: int, wrong: int, total: int, streak: int) -> str:
        if total == 0:
            return "미학습"
        if wrong >= 1 and streak < 3:
            return "복습필요"
        if streak >= 3 or (total >= 3 and correct / total >= 0.8):
            return "암기완료"
        return "학습중"


class VocabApp(tk.Tk):
    def __init__(self, store: WorkbookStore) -> None:
        super().__init__()
        self.store = store
        self.words: list[Word] = []
        self.filtered: list[Word] = []
        self.current: Word | None = None
        self.answer_var = tk.StringVar()
        self.source_var = tk.StringVar(value="전체")
        self.search_var = tk.StringVar()
        self.mode_var = tk.StringVar(value="단어 -> 뜻")

        self.title("단어 테스트")
        self.geometry("980x660")
        self.minsize(820, 560)
        self.configure(bg="#f7f8fb")

        self.build_ui()
        self.refresh_words()

    def build_ui(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background="#f7f8fb")
        style.configure("TLabel", background="#f7f8fb", font=("Malgun Gothic", 10))
        style.configure("Title.TLabel", font=("Malgun Gothic", 18, "bold"))
        style.configure("TButton", font=("Malgun Gothic", 10), padding=8)
        style.configure("Treeview", rowheight=28, font=("Malgun Gothic", 10))
        style.configure("Treeview.Heading", font=("Malgun Gothic", 10, "bold"))

        root = ttk.Frame(self, padding=18)
        root.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(root)
        header.pack(fill=tk.X)
        ttk.Label(header, text="단어 테스트", style="Title.TLabel").pack(side=tk.LEFT)
        ttk.Button(header, text="새로고침", command=self.refresh_words).pack(side=tk.RIGHT)

        controls = ttk.Frame(root)
        controls.pack(fill=tk.X, pady=(14, 10))
        ttk.Label(controls, text="출처").pack(side=tk.LEFT)
        self.source_combo = ttk.Combobox(controls, textvariable=self.source_var, state="readonly", width=18)
        self.source_combo.pack(side=tk.LEFT, padx=(8, 14))
        self.source_combo.bind("<<ComboboxSelected>>", lambda _: self.apply_filter())

        ttk.Label(controls, text="검색").pack(side=tk.LEFT)
        search = ttk.Entry(controls, textvariable=self.search_var, width=28)
        search.pack(side=tk.LEFT, padx=(8, 14))
        search.bind("<KeyRelease>", lambda _: self.apply_filter())

        ttk.Label(controls, text="모드").pack(side=tk.LEFT)
        mode = ttk.Combobox(
            controls,
            textvariable=self.mode_var,
            values=["단어 -> 뜻", "뜻 -> 단어"],
            state="readonly",
            width=14,
        )
        mode.pack(side=tk.LEFT, padx=(8, 14))

        ttk.Button(controls, text="문제 시작", command=self.next_question).pack(side=tk.LEFT)

        body = ttk.Frame(root)
        body.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(body)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 12))
        columns = ("source", "word", "meaning", "state", "accuracy")
        self.tree = ttk.Treeview(left, columns=columns, show="headings")
        self.tree.heading("source", text="출처")
        self.tree.heading("word", text="단어")
        self.tree.heading("meaning", text="뜻")
        self.tree.heading("state", text="상태")
        self.tree.heading("accuracy", text="정답률")
        self.tree.column("source", width=90, anchor=tk.W)
        self.tree.column("word", width=150, anchor=tk.W)
        self.tree.column("meaning", width=260, anchor=tk.W)
        self.tree.column("state", width=90, anchor=tk.CENTER)
        self.tree.column("accuracy", width=80, anchor=tk.CENTER)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(left, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        quiz = ttk.Frame(body, padding=14)
        quiz.pack(side=tk.RIGHT, fill=tk.BOTH)
        ttk.Label(quiz, text="문제", style="Title.TLabel").pack(anchor=tk.W)
        self.question_label = ttk.Label(quiz, text="문제 시작을 누르세요.", wraplength=310, font=("Malgun Gothic", 16, "bold"))
        self.question_label.pack(fill=tk.X, pady=(18, 12))
        self.meta_label = ttk.Label(quiz, text="", wraplength=310)
        self.meta_label.pack(fill=tk.X, pady=(0, 12))

        self.choice_frame = ttk.Frame(quiz)
        self.choice_frame.pack(fill=tk.X)
        self.choice_buttons: list[ttk.Radiobutton] = []
        for _ in range(4):
            button = ttk.Radiobutton(self.choice_frame, variable=self.answer_var, value="", text="")
            button.pack(anchor=tk.W, fill=tk.X, pady=5)
            self.choice_buttons.append(button)

        actions = ttk.Frame(quiz)
        actions.pack(fill=tk.X, pady=(18, 0))
        ttk.Button(actions, text="정답 확인", command=self.submit_answer).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(actions, text="다음 문제", command=self.next_question).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))

        self.result_label = ttk.Label(quiz, text="", wraplength=310, font=("Malgun Gothic", 11, "bold"))
        self.result_label.pack(fill=tk.X, pady=(18, 0))
        self.example_label = ttk.Label(quiz, text="", wraplength=310)
        self.example_label.pack(fill=tk.X, pady=(10, 0))

    def refresh_words(self) -> None:
        try:
            self.words = self.store.list_words()
        except Exception as exc:
            messagebox.showerror("단어 테스트", f"단어DB.xlsx를 읽을 수 없습니다.\n\n{exc}")
            return

        sources = ["전체"] + sorted({word.source for word in self.words if word.source})
        self.source_combo["values"] = sources
        if self.source_var.get() not in sources:
            self.source_var.set("전체")
        self.apply_filter()

    def apply_filter(self) -> None:
        source = self.source_var.get()
        keyword = self.search_var.get().strip().lower()
        self.filtered = []
        for word in self.words:
            if source != "전체" and word.source != source:
                continue
            searchable = " ".join([word.word, word.meaning, word.example, word.note]).lower()
            if keyword and keyword not in searchable:
                continue
            self.filtered.append(word)

        self.tree.delete(*self.tree.get_children())
        for word in self.filtered:
            self.tree.insert(
                "",
                tk.END,
                values=(word.source, word.word, word.meaning, word.memory_state, word.accuracy),
            )

    def next_question(self) -> None:
        if not self.filtered:
            messagebox.showinfo("단어 테스트", "출제할 단어가 없습니다.")
            return

        self.current = random.choice(self.filtered)
        reverse = self.mode_var.get() == "뜻 -> 단어"
        question = self.current.meaning if reverse else self.current.word
        answer = self.current.word if reverse else self.current.meaning
        candidates = [word.word if reverse else word.meaning for word in self.words if word.row != self.current.row]
        choices = random.sample(candidates, min(3, len(candidates)))
        choices.append(answer)
        random.shuffle(choices)

        self.answer_var.set("")
        self.question_label.configure(text=question)
        self.meta_label.configure(text=f"{self.current.source} | {self.current.memory_state} | 정답률 {self.current.accuracy}")
        self.result_label.configure(text="")
        self.example_label.configure(text="")
        for index, button in enumerate(self.choice_buttons):
            if index < len(choices):
                button.configure(text=choices[index], value=choices[index], state=tk.NORMAL)
                button.pack(anchor=tk.W, fill=tk.X, pady=5)
            else:
                button.pack_forget()

    def submit_answer(self) -> None:
        if self.current is None:
            return
        selected = self.answer_var.get()
        if not selected:
            messagebox.showinfo("단어 테스트", "답을 선택하세요.")
            return

        reverse = self.mode_var.get() == "뜻 -> 단어"
        answer = self.current.word if reverse else self.current.meaning
        is_correct = selected.strip() == answer.strip()
        self.store.save_result(self.current, is_correct)
        self.result_label.configure(text=("정답입니다." if is_correct else f"오답입니다. 정답: {answer}"))
        example = self.current.example
        if self.current.example_meaning:
            example = f"{example}\n{self.current.example_meaning}" if example else self.current.example_meaning
        self.example_label.configure(text=example)
        self.refresh_words()


def app_root() -> Path:
    cwd = Path.cwd()
    if (cwd / "단어DB.xlsx").exists():
        return Path.cwd()
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def data_root() -> Path:
    base = os.environ.get("LOCALAPPDATA")
    root = Path(base) / "VocabCard" if base else Path.home() / "AppData" / "Local" / "VocabCard"
    root.mkdir(parents=True, exist_ok=True)
    return root


if __name__ == "__main__":
    root = app_root()
    app = VocabApp(WorkbookStore(data_root() / "vocab.db", root / "단어DB.xlsx"))
    app.mainloop()
