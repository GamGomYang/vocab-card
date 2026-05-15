from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = ROOT_DIR / "단어DB.xlsx"

HEADERS = [
    "출처",
    "단어",
    "뜻",
    "예문",
    "예문뜻",
    "비고",
    "정답횟수",
    "오답횟수",
    "총시도",
    "최근결과",
    "최근테스트일",
    "연속정답",
    "암기상태",
]

NUMERIC_COLUMNS = {"정답횟수", "오답횟수", "총시도", "연속정답"}


@dataclass(frozen=True)
class VocabWord:
    id: int
    source: str
    word: str
    meaning: str
    example: str
    example_meaning: str
    note: str
    correct_count: int
    wrong_count: int
    total_count: int
    last_result: str
    last_tested_at: str
    streak_correct: int
    memory_state: str

    @property
    def accuracy(self) -> float:
        if self.total_count == 0:
            return 0.0
        return round((self.correct_count / self.total_count) * 100, 1)

    def to_api(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "source": self.source,
            "word": self.word,
            "meaning": self.meaning,
            "example": self.example,
            "exampleMeaning": self.example_meaning,
            "note": self.note,
            "correctCount": self.correct_count,
            "wrongCount": self.wrong_count,
            "totalCount": self.total_count,
            "lastResult": self.last_result,
            "lastTestedAt": self.last_tested_at,
            "streakCorrect": self.streak_correct,
            "memoryState": self.memory_state,
            "accuracy": self.accuracy,
        }


class ExcelRepository:
    def __init__(self, path: Path = DEFAULT_EXCEL_PATH) -> None:
        self.path = path
        self._lock = Lock()
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        if not self.path.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Words"
            sheet.append(HEADERS)
            workbook.save(self.path)
            return

        workbook = load_workbook(self.path)
        sheet = workbook.active
        changed = self._ensure_headers(sheet)
        if changed:
            workbook.save(self.path)

    def _ensure_headers(self, sheet: Worksheet) -> bool:
        existing = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        changed = False
        for header in HEADERS:
            if header not in existing:
                sheet.cell(row=1, column=len(existing) + 1, value=header)
                existing.append(header)
                changed = True

        header_map = self._header_map(sheet)
        for row in range(2, sheet.max_row + 1):
            for column_name in NUMERIC_COLUMNS:
                cell = sheet.cell(row=row, column=header_map[column_name])
                if cell.value in (None, ""):
                    cell.value = 0
                    changed = True
            if sheet.cell(row=row, column=header_map["암기상태"]).value in (None, ""):
                sheet.cell(row=row, column=header_map["암기상태"]).value = "미학습"
                changed = True
        return changed

    def _header_map(self, sheet: Worksheet) -> dict[str, int]:
        return {
            str(sheet.cell(row=1, column=index).value): index
            for index in range(1, sheet.max_column + 1)
            if sheet.cell(row=1, column=index).value
        }

    @staticmethod
    def _clean(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _int(value: Any) -> int:
        if value in (None, ""):
            return 0
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    def list_words(self) -> list[VocabWord]:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook.active
            self._ensure_headers(sheet)
            headers = self._header_map(sheet)
            words: list[VocabWord] = []

            for row in range(2, sheet.max_row + 1):
                word = self._clean(sheet.cell(row=row, column=headers["단어"]).value)
                meaning = self._clean(sheet.cell(row=row, column=headers["뜻"]).value)
                if not word or not meaning:
                    continue

                words.append(
                    VocabWord(
                        id=row - 1,
                        source=self._clean(sheet.cell(row=row, column=headers["출처"]).value),
                        word=word,
                        meaning=meaning,
                        example=self._clean(sheet.cell(row=row, column=headers["예문"]).value),
                        example_meaning=self._clean(sheet.cell(row=row, column=headers["예문뜻"]).value),
                        note=self._clean(sheet.cell(row=row, column=headers["비고"]).value),
                        correct_count=self._int(sheet.cell(row=row, column=headers["정답횟수"]).value),
                        wrong_count=self._int(sheet.cell(row=row, column=headers["오답횟수"]).value),
                        total_count=self._int(sheet.cell(row=row, column=headers["총시도"]).value),
                        last_result=self._clean(sheet.cell(row=row, column=headers["최근결과"]).value),
                        last_tested_at=self._clean(sheet.cell(row=row, column=headers["최근테스트일"]).value),
                        streak_correct=self._int(sheet.cell(row=row, column=headers["연속정답"]).value),
                        memory_state=self._clean(sheet.cell(row=row, column=headers["암기상태"]).value) or "미학습",
                    )
                )
            return words

    def get_word(self, word_id: int) -> VocabWord | None:
        return next((word for word in self.list_words() if word.id == word_id), None)

    def update_answer(self, word_id: int, is_correct: bool) -> VocabWord:
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook.active
            self._ensure_headers(sheet)
            headers = self._header_map(sheet)
            row = word_id + 1

            if row < 2 or row > sheet.max_row:
                raise KeyError(f"Word id {word_id} not found")

            correct_count = self._int(sheet.cell(row=row, column=headers["정답횟수"]).value)
            wrong_count = self._int(sheet.cell(row=row, column=headers["오답횟수"]).value)
            total_count = self._int(sheet.cell(row=row, column=headers["총시도"]).value)
            streak_correct = self._int(sheet.cell(row=row, column=headers["연속정답"]).value)

            if is_correct:
                correct_count += 1
                streak_correct += 1
                last_result = "정답"
            else:
                wrong_count += 1
                streak_correct = 0
                last_result = "오답"

            total_count += 1
            memory_state = self._memory_state(correct_count, wrong_count, total_count, streak_correct)

            sheet.cell(row=row, column=headers["정답횟수"]).value = correct_count
            sheet.cell(row=row, column=headers["오답횟수"]).value = wrong_count
            sheet.cell(row=row, column=headers["총시도"]).value = total_count
            sheet.cell(row=row, column=headers["최근결과"]).value = last_result
            sheet.cell(row=row, column=headers["최근테스트일"]).value = datetime.now().strftime("%Y-%m-%d %H:%M")
            sheet.cell(row=row, column=headers["연속정답"]).value = streak_correct
            sheet.cell(row=row, column=headers["암기상태"]).value = memory_state

            workbook.save(self.path)

        updated = self.get_word(word_id)
        if updated is None:
            raise KeyError(f"Word id {word_id} not found")
        return updated

    @staticmethod
    def _memory_state(correct_count: int, wrong_count: int, total_count: int, streak_correct: int) -> str:
        if total_count == 0:
            return "미학습"
        if wrong_count >= 1 and streak_correct < 3:
            return "복습필요"
        if streak_correct >= 3 or (total_count >= 3 and correct_count / total_count >= 0.8):
            return "암기완료"
        return "학습중"
